import streamlit as st


from openpyxl import Workbook
workbook = Workbook()
sheet = workbook.active
sheet["A1"] = "TIME/NAME"
storagelocation = "/app/detailing.xls" #"/Users/weiyushit/OneDrive/Github stuff/teststreamlit/detailing.xls" #"/home/ec2-user/teststreamlit/detailing.xlsx" #for aws 

from datetime import datetime
import time
from datetime import date
import pytz
tz = pytz.timezone('Asia/Singapore')
now = datetime.now(tz)
today = str(date.today())
current_time = now.strftime("%H:%M:%S")
st.header('Welcome to 606 CPC Detailing Planner')
st.subheader("It is " + today + " " +  current_time + "hrs")



import base64
import os
import json
import pickle
import uuid
import re
import pandas as pd
def download_link(object_to_download, download_filename, button_text, pickle_it=False):
    """
    Generates a link to download the given object_to_download.

    Params:
    ------
    object_to_download:  The object to be downloaded.
    download_filename (str): filename and extension of file. e.g. mydata.csv,
    some_txt_output.txt download_link_text (str): Text to display for download
    link.
    button_text (str): Text to display on download button (e.g. 'click here to download file')
    pickle_it (bool): If True, pickle file.

    Returns:
    -------
    (str): the anchor tag to download object_to_download

    Examples:
    --------
    download_link(your_df, 'YOUR_DF.csv', 'Click to download data!')
    download_link(your_str, 'YOUR_STRING.txt', 'Click to download text!')

    """
    if pickle_it:
        try:
            object_to_download = pickle.dumps(object_to_download)
        except pickle.PicklingError as e:
            st.write(e)
            return None
    else:
        if isinstance(object_to_download, bytes):
            pass
        elif isinstance(object_to_download, pd.DataFrame):
            object_to_download = pd.read_excel(storagelocation).to_csv()
            print("Ran workbook.save and converted it to excel")
            # object_to_download = object_to_download.to_csv(index=False)
        # Try JSON encode for everything else
        else:
            object_to_download = json.dumps(object_to_download)
    try:
        # some strings <-> bytes conversions necessary here
        b64 = base64.b64encode(object_to_download.encode()).decode()
    except AttributeError as e:
        b64 = base64.b64encode(object_to_download).decode()

    button_uuid = str(uuid.uuid4()).replace('-', '')
    button_id = re.sub('\d+', '', button_uuid)

    custom_css = f""" 
        <style>
            #{button_id} {{
                background-color: rgb(255, 255, 255);
                color: rgb(38, 39, 48);
                padding: 0.25em 0.38em;
                position: relative;
                text-decoration: none;
                border-radius: 4px;
                border-width: 1px;
                border-style: solid;
                border-color: rgb(230, 234, 241);
                border-image: initial;
            }} 
            #{button_id}:hover {{
                border-color: rgb(246, 51, 102);
                color: rgb(246, 51, 102);
            }}
            #{button_id}:active {{
                box-shadow: none;
                background-color: rgb(246, 51, 102);
                color: white;
                }}
        </style> """
    dl_link = custom_css + f'<a download="{download_filename}" id="{button_id}" href="data:file/txt;base64,{b64}">{button_text}</a><br></br>'
    return dl_link



#defaults
add_duty = []

noofdays = int(st.sidebar.slider("Days of mount", min_value=2, max_value=3, value=2, step=1))
if noofdays == 3:
    status = "weekend"
else:
    status = "weekday"

platoon = st.sidebar.selectbox(
    "Mounting Shift",
    ("Shift 1","Shift 2"),
)

#recalled personnel:
st.sidebar.subheader("Recalled Personnel")
recalled1 = st.sidebar.text_input("Name 1").upper()
recalled2 = st.sidebar.text_input("Name 2").upper()
recalled_personnel = []
if recalled1:
    recalled_personnel.append(recalled1)
if recalled2:
    recalled_personnel.append(recalled2)



if platoon == "Shift 1":
    batch0 = ["DERRICK","KEI FUNG","DYLAN PANG","BRANSON LIM","ANDRE"] 
    batch1 = ["SHAO CONG","WINSTON","AMOS","HAN TAT","MING SHENG","ZIHE","ZI KANG","BENJAMIN"]
    batch2 = [] 
    batch3 = []
    acf = []#"MING SHENG","ZIHE","ZI KANG","BENJAMIN"] #eg force
    batch4 = []
    batch5 = []
    stayout = []
else:
    batch0 = ["JOWELL","CLARENCE"] 
    batch1 = ["RAKESH","GAVIN","YONG CHENG","KOK CHUN","WEI HAN","BING HUI","CHEE SOON","YING HAO","YASHWIT","RYAN CHIANG","ALVIN SEAH",]
    batch2 = []
    batch3 = []
    acf = []#"YING HAO","YASHWIT","RYAN CHIANG","ALVIN SEAH",] #eg force
    batch4 = []
    batch5 = []
    stayout = []

team = batch0 + batch1 + batch2 + batch3 + batch4 + batch5 + stayout + recalled_personnel + ["COUNTER"]
present = [] 
for name in team:
    if st.sidebar.checkbox(name,value=True):
        present.append(name) #present now contains "COUNTER"


##

import random
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill, colors, Fill
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule

# storagelocation = "hello_world.xlsx" #for mobile compiler

#Off and Leave system #don't use if gone for whole mount. can just remove from array above
whoandwhenpresent = { #[name:day_of_mount_gone]
    'Aaron': [1,2],
}

workbook = Workbook()
sheet = workbook.active
sheet["A1"] = "TIME/NAME"
alphabets = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]

#initialise
def assigning(row, duty):
    randomperson = random.randint(2, peoplepresent+1) #+1 to this value
    if row == 2 or (sheet.cell(row = row, column = randomperson).value == None and sheet.cell(row = row-2, column = randomperson).value == None):
        sheet.cell(row = row, column = randomperson).value = duty
        sheet.cell(row = row+1, column = randomperson).value = duty
        sheet.cell(row = row+2, column = randomperson).value = duty
        sheet.cell(row = row+3, column = randomperson).value = duty
    else:
        assigning(row, duty)

def assigningpeak(row, duty): #4hourblock 1 peak 1 non-peak
    randomperson = random.randint(2, peoplepresent+1)
    if sheet.cell(row = row, column = randomperson).value == None and sheet.cell(row = row-1, column = randomperson).value == None and (duty != "GPMG"):
        sheet.cell(row = row, column = randomperson).value = duty
        sheet.cell(row = row+1, column = randomperson).value = duty
        sheet.cell(row = row+2, column = randomperson).value = duty
        sheet.cell(row = row+3, column = randomperson).value = duty
        colourthisrow(i,"ff0000")
        colourthisrow(i+1,"ff0000")
    elif (sheet.cell(row = row, column = randomperson).value == None and sheet.cell(row = row-2, column = randomperson).value == None) and (duty == "GPMG"):
        sheet.cell(row = row, column = randomperson).value = duty
        sheet.cell(row = row+1, column = randomperson).value = duty
        colourthisrow(i,"ff0000")
        colourthisrow(i+1,"ff0000")
    else:
        assigningpeak(row, duty)

def assigningafterpeak(counter,duty):
    if sheet.cell(row = i, column = counter).value == duty:
        sheet.cell(row = i+1, column = counter).value = duty
    else:
        counter += 1
        assigningafterpeak(counter,duty)

def countcellstoleft(row): #counter
    answer = 0
    for i in range(2, peoplepresent+2): #+2 cuz range skips the last one
        if sheet.cell(row= row, column= i).value != None:
            answer += 1
    return int(answer)

def hourscounter():
    for i in range(2, peoplepresent+2):
        counterhour = 0
        for row in range(2, totalrows):
            if sheet.cell(row=row, column = i).value != None and sheet.cell(row=row, column = i).value != "F":
                counterhour += 1
        sheet.cell(row=hoursrow, column = i).value = counterhour

def xinjiaolaojiaosystem():
    hoursranking = {}
    for i in range(2, peoplepresent+2):
        hours = sheet.cell(row=hoursrow, column = i).value
        hoursranking[i] = hours
    hoursranking = {k: v for k, v in sorted(hoursranking.items(), key=lambda item: item[1])}
    columnorder = list(hoursranking.keys()) 
    for i in range(len(team)-1):
        sheet.cell(row=1, column=columnorder[i]).value = team[i]
    return

def add_flags():
    for i in range(2,totalrows):
        if sheet.cell(row=i, column = 1).value in ["1800","0600"]: #give flag to person who has duty next
            print("Not Done, working on flag")



#set up time
row = 2
timedefault = ["1100","1200","1300","1400","1500","1600","1700","1800","1900","2000","2100","2200","2300","0000","0100","0200","0300","0400","0500","0600","0700","0800","0900","1000"]
# timedefault = ["1100-1300", "1300-1500","1500-1700", "1700-1900","1900-2100", "2100-2300", "2300-0100","0100-0300", "0300-0500","0500-0700","0700-0900","0900-1100"]
times = noofdays * timedefault
for timeblock in times:
    sheet.cell(row = row, column = 1).value = timeblock
    row += 1
totalrows = row
hoursrow = totalrows #dunnid to add one more because final interation of timeblock already adds 1 more.
sheet.cell(row = hoursrow, column = 1).value = "TOTAL"

#set up humans, minimum 19
team = present
peoplepresent = len(present)-1 #REMOVE 1 CUZ OF COUNTER
print("Number of people present is {}".format(str(peoplepresent))) 
column = 2
#set up column headers
for name in team:
    sheet.cell(row = 1, column = column).value = name
    column += 1

def process_stayout_first():
    stayoutcolumn = len(team)
    for i in range(2, totalrows):
        if status == "weekday" and (sheet.cell(row=i, column=1).value in ["1700-1900","1900-2100","2100-2300","2300-0100","0100-0300","0300-0500","0500-0700"]):
            sheet.cell(row=i, column=stayoutcolumn).value = "STAYOUT"
        if status == "weekend":
            if i >= 5 and i <= totalrows-3:
                sheet.cell(row=i, column=stayoutcolumn).value = "STAYOUT"
#process_stayout_first()

# # # #def leaveandoffs(who,whichdays):
# # # for name in team:
# # #     if name in list(whoandwhenpresent.keys()):
# # #         for daynumber in whoandwhenpresent[name]:
# # #             print(name + " is on leave for day " + str(daynumber))
# # #             if daynumber == 1:
# # #                 sheet.cell(row="person's row",column="person's column").value = "OFF/LL"

#colour coding
def colourthisrow(row,colour):
    for i in range(0,len(team)+1): #+1 cuz need account for time and counter column
        columncoordinate = alphabets[i]
        cellcoordinate = columncoordinate + str(row)
        cell = sheet[cellcoordinate]
        cell.fill = PatternFill(start_color=colour, end_color=colour, fill_type = "solid")



#initialise duties
non_peak = ["SSVC","SCBT","XCBT","XSVC"]
peak = ["SSVC","SCBT","XCBT","XSVC"]
silent = [e for e in non_peak if e not in ('XSVC', 'XCBT')]

#if heightened measure add gpmg and footprowl
checkheightened = st.sidebar.checkbox("Heightened")
if checkheightened:
    #add GPMG and Foot Prowl
    add_duty = ["GPMG"]
    peak.extend(add_duty)  
    def add_prowls():
        #find rows that are 0100, 0500, 1900, 2300hrs, check if they have duty above, then plot detailing
        print("Not Done, working on flag")

non_peak_hours = ["1100","1200","1300","1400","1500","1600","1700","1800","0700","0800","0900","1000"]
peak_hours = ["0700","0800","1600","1700","1800"]
silent_hours = ["1900","2000","2100","2200","2300","0000","0100","0200","0300","0400","0500","0600"]
#assign dutytypes to hours
#nonpeak = 7, peak = 10, silent = 5
row = 2 #reset row again
print("planning....")


if status == "weekday":
    for i in range(2, totalrows):
        if (sheet.cell(row= i, column = 1).value in ["1100","1500"]): #if non_peak on normal hours
            for duty in non_peak:
                assigning(i, duty)
            #if cell is empty (leave, off, MA etc) then put into random
        if (sheet.cell(row= i, column = 1).value in "0700"): #peakhour
            for duty in peak:
                assigningpeak(i,duty)
            counter = 1 #function below for adding non-peak for 0900-1100
            for duty in non_peak:
                assigningafterpeak(counter,duty)
        if (sheet.cell(row= i, column = 1).value in ["1900","2300","0300"]):
            colourthisrow(i,"808080")
            colourthisrow(i+1,"808080")
            colourthisrow(i+2,"808080")
            colourthisrow(i+3,"808080")
            for duty in silent:
                assigning(i,duty)
        sheet.cell(row=i, column= peoplepresent+2).value = int(countcellstoleft(i))
elif status == "weekend":
    for i in range(2, totalrows):
        if (sheet.cell(row= i, column = 1).value in ["0700","1100","1500"]): #if non_peak on normal hours
            if i<=8 or i>=70:
                for duty in non_peak:
                    assigning(i, duty)
                #if cell is empty (leave, off, MA etc) then put into random
            else:
                colourthisrow(i,"808080")
                colourthisrow(i+1,"808080")
                colourthisrow(i+2,"808080")
                colourthisrow(i+3,"808080")
                for duty in silent:
                    assigning(i,duty)
        if (sheet.cell(row= i, column = 1).value in ["0700"]):
            if i>=70:
                colourthisrow(i,"ff0000")
                for duty in peak:
                    assigningpeak(i,duty)
        if (sheet.cell(row= i, column = 1).value in ["1900","2300","0300"]):
            colourthisrow(i,"808080")
            colourthisrow(i+1,"808080")
            colourthisrow(i+2,"808080")
            colourthisrow(i+3,"808080")
            for duty in silent:
                assigning(i,duty)
        sheet.cell(row=i, column= peoplepresent+2).value = int(countcellstoleft(i))



hourscounter()
#xinjiaolaojiaosystem()
# add_prowls()
add_flags()
print("Done.")
workbook.save(filename=storagelocation)


###
#displaying excel sheet converted to pandas
import pandas as pd
if True:
    df = pd.read_excel(storagelocation)
    st.dataframe(df.fillna(" "))
    
    #st.table(df.fillna(" "))


st.button("Rerun")

if st.button("Export this"):
    st.write("Exporting...")
    #function to assign excelfile to a variable, then provide download link for it.
    df = pd.read_excel(storagelocation)
    #Exporting/Downloading excel sheet
    download_button_str = download_link(df, "Completed Detailing.xls", 'Click here to download Completed Detailing.csv', pickle_it=False)
    st.markdown(download_button_str, unsafe_allow_html=True)

##TODo
'''

FUTURE IMPROVEMENTS
1. Add flag function
2. Fix weekend planning
3. Add indication of version
4. Figure how to add timing for people who took off, mc, leave not for whole mount, and recalled personnel
5. Auto-balancing of hours algorithm


LONG TERM
A. make it easy to edit name list for future generations
B. add duty posts
(Consider creating another streamlit app to edit code of repo and deploy wow)
'''
